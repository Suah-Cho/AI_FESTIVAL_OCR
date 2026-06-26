"""엑셀 헤더 탐색 및 셀 색칠 유틸리티.

- 헤더 행 위치를 고정하지 않고, 입력한 열 이름들이 가장 많이 발견되는 행을
  헤더 행으로 판단한다.
- 검증 대상 열의 셀 자체에 색을 칠한다.
"""

from __future__ import annotations

import re
import unicodedata
from dataclasses import dataclass, field
from datetime import date, datetime, time
from typing import Any, Optional

from openpyxl.cell.cell import Cell
from openpyxl.styles import PatternFill
from openpyxl.utils.datetime import from_excel
from openpyxl.worksheet.worksheet import Worksheet

from app.services.field_aliases import expanded_header_norms, find_column_index

# 색상 정의 (요구사항)
FILL_MISMATCH = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # 빨강
FILL_NEEDS_CHECK = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # 노랑
FILL_MATCH = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # 초록
FILL_EXCLUDED = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")  # 회색

# 헤더 탐색 시 살펴볼 최대 행 수
MAX_HEADER_SCAN_ROWS = 30


def _excel_date_separator(number_format: str) -> str:
    """셀 서식에서 날짜 구분자(., /, -)를 추정한다."""
    fmt = number_format or ""
    if "/" in fmt:
        return "/"
    if re.search(r"yyyy[\-.]mm", fmt, re.I):
        return "-" if "-" in fmt else "."
    if "\\." in fmt or re.search(r"mm\.dd", fmt, re.I):
        return "."
    return "-"


def _looks_like_date_only_format(number_format: str) -> bool:
    fmt = (number_format or "").lower()
    if not fmt or fmt == "general":
        return False
    has_date = "y" in fmt and "d" in fmt
    has_time = "h" in fmt or "ss" in fmt or ":mm" in fmt
    return has_date and not has_time


def format_cell_display(value: Any, number_format: Optional[str] = None) -> str:
    """openpyxl 셀 값을 화면·비교용 문자열로 변환 (날짜의 00:00:00 제거)."""
    if value is None:
        return ""

    fmt = str(number_format or "General")

    if isinstance(value, (int, float)) and _looks_like_date_only_format(fmt):
        try:
            value = from_excel(value)
        except (ValueError, OverflowError, TypeError):
            pass

    if isinstance(value, datetime):
        if value.time() == time.min:
            sep = _excel_date_separator(fmt)
            return f"{value.year:04d}{sep}{value.month:02d}{sep}{value.day:02d}"
        return value.strftime("%Y-%m-%d %H:%M:%S")

    if isinstance(value, date):
        sep = _excel_date_separator(fmt)
        return f"{value.year:04d}{sep}{value.month:02d}{sep}{value.day:02d}"

    return unicodedata.normalize("NFKC", str(value)).strip()


def cell_display_text(cell: Cell) -> str:
    """워크시트 셀 하나를 화면에 보여줄 문자열로 변환한다."""
    return format_cell_display(cell.value, cell.number_format)


def _norm_header(text: Optional[str]) -> str:
    """헤더 비교용 정규화: 유니코드 정규화 + 공백 제거."""
    if text is None:
        return ""
    return unicodedata.normalize("NFKC", str(text)).strip().replace(" ", "")


@dataclass
class HeaderLayout:
    """탐색된 헤더 행 정보."""

    header_row: int
    column_index: dict[str, int] = field(default_factory=dict)  # 정규화이름 -> 열(1-base)
    raw_headers: dict[int, str] = field(default_factory=dict)  # 열 -> 원본 헤더 텍스트

    def find_column(self, name: str) -> Optional[int]:
        return find_column_index(self.column_index, name)


def detect_header_row(ws: Worksheet, candidate_names: list[str]) -> Optional[HeaderLayout]:
    """후보 열 이름들이 가장 많이 매칭되는 행을 헤더 행으로 판단한다."""
    wanted = expanded_header_norms(candidate_names)
    if not wanted:
        return None

    best_row = None
    best_hits = 0
    scan_limit = min(ws.max_row, MAX_HEADER_SCAN_ROWS)
    for row_idx in range(1, scan_limit + 1):
        hits = sum(1 for cell in ws[row_idx] if _norm_header(cell.value) in wanted)
        if hits > best_hits:
            best_hits = hits
            best_row = row_idx

    if best_row is None or best_hits == 0:
        return None

    layout = HeaderLayout(header_row=best_row)
    for cell in ws[best_row]:
        norm = _norm_header(cell.value)
        if norm and norm not in layout.column_index:
            layout.column_index[norm] = cell.column
            layout.raw_headers[cell.column] = cell_display_text(cell)
    return layout


def resolve_columns(
    layout: HeaderLayout, target_columns: list[str]
) -> tuple[dict[str, int], list[str]]:
    """입력한 열 이름들을 실제 열 인덱스로 매핑한다.

    반환: (이름->열인덱스 매핑, 못 찾은 이름 목록)
    """
    resolved: dict[str, int] = {}
    missing: list[str] = []
    for name in target_columns:
        col = layout.find_column(name)
        if col is None:
            missing.append(name)
        else:
            resolved[name] = col
    return resolved, missing


def append_result_column(ws: Worksheet, layout: HeaderLayout, title: str = "검증결과") -> int:
    """헤더 행 맨 끝에 결과 열을 추가하고 그 열 인덱스를 반환한다.

    이미 같은 제목 열이 있으면 그 열을 재사용한다.
    """
    existing = layout.find_column(title)
    if existing is not None:
        return existing
    new_col = ws.max_column + 1
    ws.cell(row=layout.header_row, column=new_col, value=title)
    return new_col


def fill_cell(ws: Worksheet, row: int, col: int, fill: PatternFill) -> None:
    ws.cell(row=row, column=col).fill = fill


def clear_cell_fill(ws: Worksheet, row: int, col: int) -> None:
    """셀 배경색을 제거한다 (수정 반영 후 다운로드용)."""
    ws.cell(row=row, column=col).fill = PatternFill()
