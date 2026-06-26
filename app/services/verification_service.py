"""핵심 검증 로직.

흐름: (업로드 파일 정리) → 엑셀 읽기 → 헤더에서 입력 열 찾기 → 폴더 매칭
      → GPT 추출 → 정규화/판정 → 셀 색칠/요약 → 결과 엑셀 저장

라우터는 ``verify_uploads`` (또는 ``verify_zip``) 만 호출하면 된다.
"""

from __future__ import annotations

import logging
import os
import tempfile
import zipfile
from dataclasses import dataclass, field
from pathlib import Path
from typing import Optional

from openpyxl import load_workbook

from app.core.config import get_settings
from app.schemas.verification import VerifyResult
from app.services import excel_service as xl
from app.services.extraction_service import (
    extract_fields_from_documents,
    has_documents,
)
from app.services.field_aliases import (
    all_location_header_names,
    is_rrn_field,
    is_valid_rrn_format,
    resolve_location_columns,
    split_doc_path,
)
from app.services.field_extraction_service import BATCH_PER_FOLDER, group_document_units
from app.services.normalization import (
    EXCLUDED,
    MATCH,
    MISMATCH,
    NEEDS_CHECK,
    compare_values,
)

logger = logging.getLogger(__name__)


class VerificationError(Exception):
    """검증 진행이 불가능한 사용자 입력/데이터 오류."""


def _find_xlsx(extract_dir: Path) -> Optional[Path]:
    """디렉터리에서 .xlsx 파일을 찾는다 (임시/숨김 파일 제외)."""
    candidates = [
        p
        for p in extract_dir.rglob("*.xlsx")
        if not p.name.startswith("~$") and "__MACOSX" not in p.parts
    ]
    return candidates[0] if candidates else None


def _cell_text(value) -> str:
    return "" if value is None else str(value).strip()


def _status_to_fill(status: str):
    return {
        MATCH: xl.FILL_MATCH,
        MISMATCH: xl.FILL_MISMATCH,
        NEEDS_CHECK: xl.FILL_NEEDS_CHECK,
        EXCLUDED: xl.FILL_EXCLUDED,
    }[status]


def _find_doc_folder(base_dir: Path, folder_val: str, file_no_val: str) -> Optional[Path]:
    """``{파일폴더}/{파일번호}`` 폴더를 찾는다.

    base_dir 기준으로 우선 탐색하고, 못 찾으면 하위 전체에서 같은 구조를 찾는다
    (ZIP 내부 폴더 깊이가 달라도 매칭되도록).
    """
    folder_val = folder_val.strip()
    file_no_val = file_no_val.strip()

    direct = base_dir / folder_val / file_no_val
    if direct.is_dir():
        return direct

    for candidate in base_dir.rglob(file_no_val):
        if candidate.is_dir() and candidate.parent.name == folder_val:
            return candidate

    for candidate in base_dir.rglob(file_no_val):
        if candidate.is_dir():
            return candidate

    return None


def _list_doc_units(doc_base_dir: Path) -> list[tuple[str, str, str]]:
    """ZIP 서류 구조에서 (파일폴더값, 파일번호값, 폴더경로) 목록을 만든다."""
    units = group_document_units(doc_base_dir, BATCH_PER_FOLDER)
    result: list[tuple[str, str, str]] = []
    for unit in units:
        label = unit.label.replace("\\", "/")
        if "/" in label:
            folder_val, file_no_val = split_doc_path(label)
        else:
            folder_val, file_no_val = "", label
        result.append((folder_val, file_no_val, unit.folder_path))
    return result


def _format_header_list(raw_headers: dict[int, str]) -> str:
    ordered = [raw_headers[c] for c in sorted(raw_headers)]
    return ", ".join(ordered) if ordered else "(헤더 없음)"


def _row_expected_values(prepared: PreparedSheet, plan: RowPlan) -> dict[str, str]:
    """행의 엑셀 기준값(검증 대상 열)을 읽는다."""
    return {
        name: _cell_text(prepared.ws.cell(row=plan.row_index, column=col).value)
        for name, col in plan.expected.items()
    }


def _display_value(value: str) -> str:
    text = _cell_text(value)
    return text if text else "(없음)"


def _format_field_detail(
    name: str,
    status: str,
    expected: str,
    extracted: str,
) -> str:
    """불일치·확인필요 항목 한 줄 설명."""
    exp = _display_value(expected)
    ext = _display_value(extracted)
    if status == MISMATCH:
        return f"{name} 불일치 — 기준「{exp}」 서류「{ext}」"
    if status == NEEDS_CHECK:
        if not _cell_text(extracted):
            return f"{name} 확인필요 — 서류에서 못 찾음 (기준「{exp}」)"
        return f"{name} 확인필요 — 기준「{exp}」 서류「{ext}」"
    if status == EXCLUDED:
        return f"{name} 제외"
    return ""


def _build_result_text(
    statuses: dict[str, str],
    expected: dict[str, str],
    extracted: dict[str, str],
) -> str:
    """검증결과 열·화면용 요약 문구."""
    if all(status == MATCH for status in statuses.values()):
        return "OK"
    parts = [
        _format_field_detail(
            name,
            status,
            expected.get(name, ""),
            extracted.get(name, ""),
        )
        for name, status in statuses.items()
        if status != MATCH
    ]
    return " | ".join(p for p in parts if p)


@dataclass
class RowPlan:
    """검증 대상 한 행에 대한 사전 계산 정보."""

    row_index: int  # 엑셀 행 번호 (1-base)
    folder_val: str
    file_no_val: str
    doc_folder: Optional[str]  # 매칭된 서류 폴더 경로 (없으면 None)
    excluded: bool  # 파일폴더/파일번호가 비어 검증 제외 대상인지
    expected: dict[str, int]  # 열이름 -> 열 인덱스 (1-base)


@dataclass
class RowResult:
    """한 행 검증 결과."""

    row_index: int
    statuses: dict[str, str]  # 열이름 -> 판정(MATCH/MISMATCH/...)
    result_text: str
    kind: str = "verified"  # excluded / no_docs / verified / error
    expected: dict[str, str] = field(default_factory=dict)  # 열이름 -> 엑셀 기준값
    extracted: dict[str, str] = field(default_factory=dict)  # 열이름 -> 서류 추출값


@dataclass
class PreparedSheet:
    """엑셀을 읽어 검증 준비를 마친 상태. 행 단위로 병렬 검증할 수 있다."""

    wb: object
    ws: object
    target_columns: list[str]
    resolved: dict[str, int]  # 열이름 -> 열 인덱스 (1-base)
    result_col: int
    header_row: int
    rows: list[RowPlan]
    cells: list[list[str]]  # 화면 표시용 그리드 (cells[r][c] == ws (r+1, c+1) 값)
    counts: dict[str, int] = field(
        default_factory=lambda: {MATCH: 0, MISMATCH: 0, NEEDS_CHECK: 0, EXCLUDED: 0}
    )
    output_path: Optional[str] = None

    @property
    def total_rows(self) -> int:
        return len(self.rows)

    def target_col_indices(self) -> dict[str, int]:
        return dict(self.resolved)


def prepare_sheet(
    xlsx_path: Path,
    doc_base_dir: Path,
    target_columns: list[str],
    folder_col_name: Optional[str] = None,
    file_no_col_name: Optional[str] = None,
) -> PreparedSheet:
    """엑셀을 읽고 헤더/열/행을 해석해 병렬 검증 준비 상태를 만든다.

    실제 LLM 호출은 하지 않는다. (화면에 표를 먼저 보여주기 위한 단계)
    """
    settings = get_settings()
    folder_col_name = folder_col_name or settings.default_folder_col
    file_no_col_name = file_no_col_name or settings.default_file_no_col

    target_columns = [c.strip() for c in target_columns if c and c.strip()]
    if not target_columns:
        raise VerificationError("검증할 열 이름을 한 개 이상 입력해 주세요.")

    wb = load_workbook(xlsx_path)
    ws = wb.active

    all_candidates = list(
        dict.fromkeys(
            target_columns
            + [folder_col_name, file_no_col_name]
            + all_location_header_names()
        )
    )
    layout = xl.detect_header_row(ws, all_candidates)
    if layout is None:
        raise VerificationError(
            "입력한 열 이름들을 헤더에서 찾지 못했습니다. 열 이름이 엑셀 헤더와 같은지 확인해 주세요."
        )

    resolved, missing = xl.resolve_columns(layout, target_columns)
    if missing:
        raise VerificationError(
            "다음 열 이름을 엑셀 헤더에서 찾지 못했습니다: " + ", ".join(missing)
        )

    location = resolve_location_columns(
        layout.column_index,
        layout.raw_headers,
        folder_col_name,
        file_no_col_name,
    )
    auto_units: list[tuple[str, str, str]] | None = None
    if location.mode == "auto":
        auto_units = _list_doc_units(doc_base_dir)
        if not auto_units:
            raise VerificationError(
                "엑셀에 서류 위치 열(파일폴더·파일번호·서류경로 등)이 없고, "
                "ZIP 안에서 서류 폴더도 찾지 못했습니다. "
                f"엑셀 헤더: {_format_header_list(layout.raw_headers)}"
            )

    result_col = xl.append_result_column(ws, layout, "검증결과")

    rows: list[RowPlan] = []
    auto_idx = 0
    for row_idx in range(layout.header_row + 1, ws.max_row + 1):
        folder_val = ""
        file_no_val = ""
        doc_folder: Optional[str] = None

        if location.mode == "two":
            folder_val = _cell_text(ws.cell(row=row_idx, column=location.folder_col).value)
            file_no_val = _cell_text(ws.cell(row=row_idx, column=location.file_no_col).value)
        elif location.mode == "path":
            path_val = _cell_text(ws.cell(row=row_idx, column=location.path_col).value)
            folder_val, file_no_val = split_doc_path(path_val)

        row_has_data = folder_val or file_no_val or any(
            _cell_text(ws.cell(row=row_idx, column=col).value) for col in resolved.values()
        )
        if not row_has_data:
            continue

        if location.mode == "auto":
            if auto_units and auto_idx < len(auto_units):
                folder_val, file_no_val, doc_folder = auto_units[auto_idx]
                auto_idx += 1
            else:
                rows.append(
                    RowPlan(
                        row_index=row_idx,
                        folder_val=folder_val,
                        file_no_val=file_no_val,
                        doc_folder=None,
                        excluded=True,
                        expected=dict(resolved),
                    )
                )
                continue
        elif not folder_val or not file_no_val:
            rows.append(
                RowPlan(
                    row_index=row_idx,
                    folder_val=folder_val,
                    file_no_val=file_no_val,
                    doc_folder=None,
                    excluded=True,
                    expected=dict(resolved),
                )
            )
            continue
        else:
            found = _find_doc_folder(doc_base_dir, folder_val, file_no_val)
            doc_folder = str(found) if found else None

        rows.append(
            RowPlan(
                row_index=row_idx,
                folder_val=folder_val,
                file_no_val=file_no_val,
                doc_folder=doc_folder,
                excluded=False,
                expected=dict(resolved),
            )
        )

    # 화면 표시용 그리드 (헤더에 추가된 검증결과 열까지 포함)
    cells: list[list[str]] = []
    for r in range(1, ws.max_row + 1):
        row_vals = [_cell_text(ws.cell(row=r, column=c).value) for c in range(1, ws.max_column + 1)]
        cells.append(row_vals)

    return PreparedSheet(
        wb=wb,
        ws=ws,
        target_columns=target_columns,
        resolved=resolved,
        result_col=result_col,
        header_row=layout.header_row,
        rows=rows,
        cells=cells,
    )


def verify_row(prepared: PreparedSheet, plan: RowPlan) -> RowResult:
    """한 행을 검증한다. (LLM 호출 포함 / 워크북은 건드리지 않음 → 병렬 안전)

    예외는 호출자에게 던지지 않고 '확인필요'로 처리해 스트리밍이 끊기지 않게 한다.
    """
    statuses: dict[str, str] = {}
    expected_values = _row_expected_values(prepared, plan)
    extracted_values: dict[str, str] = {}

    if plan.excluded:
        for name in prepared.target_columns:
            statuses[name] = EXCLUDED
        result_text = _build_result_text(statuses, expected_values, extracted_values)
        return RowResult(
            plan.row_index,
            statuses,
            result_text or "제외(폴더/번호 없음)",
            kind="excluded",
            expected=expected_values,
            extracted=extracted_values,
        )

    if plan.doc_folder is None or not has_documents(plan.doc_folder):
        for name in prepared.target_columns:
            statuses[name] = NEEDS_CHECK
        result_text = _build_result_text(statuses, expected_values, extracted_values)
        return RowResult(
            plan.row_index,
            statuses,
            result_text or "확인필요(서류 폴더 없음)",
            kind="no_docs",
            expected=expected_values,
            extracted=extracted_values,
        )

    logger.info(
        "행 검증 시작 row=%d folder=%s/%s fields=%s",
        plan.row_index,
        plan.folder_val,
        plan.file_no_val,
        prepared.target_columns,
    )

    try:
        extracted_raw = extract_fields_from_documents(plan.doc_folder, prepared.target_columns)
    except Exception as exc:  # GPT 실패/타임아웃 등 → 행 단위로 격리
        logger.error("행 검증 추출 오류 row=%d error=%s", plan.row_index, exc)
        for name in prepared.target_columns:
            statuses[name] = NEEDS_CHECK
        return RowResult(
            plan.row_index,
            statuses,
            f"확인필요(추출 오류: {exc})",
            kind="error",
            expected=expected_values,
            extracted=extracted_values,
        )

    for name, col in plan.expected.items():
        expected = expected_values[name]
        extracted_val = str(extracted_raw.get(name, "") or "")
        extracted_values[name] = extracted_val
        if is_rrn_field(name) and _cell_text(extracted_val) and not is_valid_rrn_format(
            extracted_val
        ):
            status = NEEDS_CHECK
        else:
            status = compare_values(expected, extracted_val)
        statuses[name] = status

    result_text = _build_result_text(statuses, expected_values, extracted_values)
    logger.info(
        "행 검증 완료 row=%d result=%s extracted=%s statuses=%s",
        plan.row_index,
        result_text,
        extracted_values,
        statuses,
    )

    return RowResult(
        plan.row_index,
        statuses,
        result_text,
        kind="verified",
        expected=expected_values,
        extracted=extracted_values,
    )


def apply_row_result(prepared: PreparedSheet, result: RowResult) -> None:
    """행 검증 결과를 워크북에 반영(셀 색칠 + 결과열 기록)하고 카운트를 누적한다.

    워크북 쓰기는 스레드 안전하지 않으므로 단일 스레드에서만 호출해야 한다.
    카운트 규칙: '제외' 행은 행당 1회, 그 외 판정은 셀(열)마다 1회 누적한다.
    """
    for name, status in result.statuses.items():
        col = prepared.resolved.get(name)
        if col is not None and status != MATCH:
            xl.fill_cell(prepared.ws, result.row_index, col, _status_to_fill(status))
        if result.kind != "excluded":
            prepared.counts[status] = prepared.counts.get(status, 0) + 1
    if result.kind == "excluded":
        prepared.counts[EXCLUDED] = prepared.counts.get(EXCLUDED, 0) + 1
    prepared.ws.cell(row=result.row_index, column=prepared.result_col, value=result.result_text)


def apply_cell_edits(prepared: PreparedSheet, edits: list[dict]) -> None:
    """화면에서 수정한 셀 값을 워크북에 반영하고 강조 색을 제거한다."""
    for edit in edits:
        row = int(edit["row"])
        col = int(edit["col"])
        value = edit.get("value", "")
        prepared.ws.cell(row=row, column=col, value=value)
        prepared.cells[row - 1][col - 1] = value
        xl.clear_cell_fill(prepared.ws, row, col)


def finalize(prepared: PreparedSheet, output_path: Optional[str] = None) -> str:
    """검증 반영이 끝난 워크북을 저장하고 경로를 반환한다."""
    if output_path is None:
        if prepared.output_path:
            output_path = prepared.output_path
        else:
            fd, output_path = tempfile.mkstemp(suffix=".xlsx", prefix="result_")
            os.close(fd)
    prepared.wb.save(output_path)
    prepared.output_path = output_path
    return output_path


def build_verify_result(prepared: PreparedSheet) -> VerifyResult:
    excluded_rows = sum(1 for r in prepared.rows if r.excluded)
    return VerifyResult(
        output_path=prepared.output_path or "",
        total_rows=prepared.total_rows,
        verified_rows=prepared.total_rows - excluded_rows,
        excluded_rows=excluded_rows,
        counts=dict(prepared.counts),
        missing_columns=[],
    )


def _verify_workbook(
    xlsx_path: Path,
    doc_base_dir: Path,
    target_columns: list[str],
    folder_col_name: Optional[str] = None,
    file_no_col_name: Optional[str] = None,
    output_path: Optional[str] = None,
) -> VerifyResult:
    """엑셀과 서류 폴더 루트를 받아 검증을 수행하는 핵심 함수.

    Args:
        xlsx_path: 검증 기준 엑셀 경로.
        doc_base_dir: ``{파일폴더}/{파일번호}`` 서류 폴더를 탐색할 루트 디렉터리.
        target_columns: 검증 대상 열 이름 목록.
        folder_col_name / file_no_col_name: 서류 위치 지정 열 이름 (None=설정 기본값).
        output_path: 결과 엑셀 저장 경로 (None=임시 파일).
    """
    prepared = prepare_sheet(
        xlsx_path=xlsx_path,
        doc_base_dir=doc_base_dir,
        target_columns=target_columns,
        folder_col_name=folder_col_name,
        file_no_col_name=file_no_col_name,
    )

    for plan in prepared.rows:
        apply_row_result(prepared, verify_row(prepared, plan))

    finalize(prepared, output_path)
    return build_verify_result(prepared)


def _extract_zip(zip_path: Path, dest: Path) -> None:
    try:
        with zipfile.ZipFile(zip_path) as zf:
            zf.extractall(dest)
    except zipfile.BadZipFile as exc:
        raise VerificationError(f"올바른 ZIP 파일이 아닙니다: {zip_path.name}") from exc


def _resolve_inputs(file_paths: list[str]) -> tuple[Path, Path]:
    """업로드 파일들(.xlsx/.zip 혼합)을 풀고 (검증 엑셀 경로, 서류 루트)를 반환한다."""
    paths = [Path(p) for p in file_paths]
    zip_paths = [p for p in paths if p.suffix.lower() == ".zip"]
    xlsx_paths = [
        p for p in paths if p.suffix.lower() == ".xlsx" and not p.name.startswith("~$")
    ]

    if not zip_paths and not xlsx_paths:
        raise VerificationError("ZIP 또는 XLSX 파일을 업로드해 주세요.")

    work = Path(tempfile.mkdtemp(prefix="verify_"))
    for zp in zip_paths:
        _extract_zip(zp, work)

    # 검증 기준 엑셀 결정: 따로 올린 xlsx 우선, 없으면 ZIP 내부 xlsx
    xlsx_path: Optional[Path] = xlsx_paths[0] if xlsx_paths else _find_xlsx(work)
    if xlsx_path is None:
        raise VerificationError(
            "검증 기준 엑셀(.xlsx)을 찾지 못했습니다. 엑셀 파일을 함께 업로드하거나 ZIP 안에 포함해 주세요."
        )

    return xlsx_path, work


def prepare_uploads(
    file_paths: list[str],
    target_columns: list[str],
    folder_col_name: Optional[str] = None,
    file_no_col_name: Optional[str] = None,
) -> PreparedSheet:
    """업로드 파일들을 풀어 검증 준비 상태(PreparedSheet)를 만든다. (LLM 호출 없음)"""
    xlsx_path, work = _resolve_inputs(file_paths)
    return prepare_sheet(
        xlsx_path=xlsx_path,
        doc_base_dir=work,
        target_columns=target_columns,
        folder_col_name=folder_col_name,
        file_no_col_name=file_no_col_name,
    )


def verify_uploads(
    file_paths: list[str],
    target_columns: list[str],
    folder_col_name: Optional[str] = None,
    file_no_col_name: Optional[str] = None,
    output_path: Optional[str] = None,
) -> VerifyResult:
    """업로드된 여러 파일(.xlsx 와 .zip 혼합)을 받아 (동기) 검증한다."""
    xlsx_path, work = _resolve_inputs(file_paths)
    return _verify_workbook(
        xlsx_path=xlsx_path,
        doc_base_dir=work,
        target_columns=target_columns,
        folder_col_name=folder_col_name,
        file_no_col_name=file_no_col_name,
        output_path=output_path,
    )


def verify_zip(
    zip_path: str,
    target_columns: list[str],
    folder_col_name: Optional[str] = None,
    file_no_col_name: Optional[str] = None,
    output_path: Optional[str] = None,
) -> VerifyResult:
    """ZIP 하나를 검증한다 (엑셀은 ZIP 내부에 포함). 하위 호환용 진입점."""
    extract_root = Path(tempfile.mkdtemp(prefix="verify_"))
    _extract_zip(Path(zip_path), extract_root)

    xlsx_path = _find_xlsx(extract_root)
    if xlsx_path is None:
        raise VerificationError("ZIP 안에서 .xlsx 파일을 찾지 못했습니다.")

    return _verify_workbook(
        xlsx_path=xlsx_path,
        doc_base_dir=xlsx_path.parent,
        target_columns=target_columns,
        folder_col_name=folder_col_name,
        file_no_col_name=file_no_col_name,
        output_path=output_path,
    )
