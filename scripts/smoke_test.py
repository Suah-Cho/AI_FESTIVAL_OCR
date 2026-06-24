"""더미 추출기로 전체 검증 흐름을 점검하는 스모크 테스트.

GPT 호출 없이 (USE_DUMMY_EXTRACTOR=1) ZIP 생성 → verify_zip → 결과 확인.
실행: python scripts/smoke_test.py
"""

from __future__ import annotations

import json
import os
import tempfile
import zipfile
from pathlib import Path

os.environ["USE_DUMMY_EXTRACTOR"] = "1"

from openpyxl import Workbook, load_workbook  # noqa: E402

from app.services.verification_service import verify_zip  # noqa: E402


def build_fixture(root: Path) -> Path:
    """테스트용 ZIP 파일을 만들어 경로를 반환한다."""
    work = root / "이마트24_2"
    work.mkdir(parents=True, exist_ok=True)

    # 엑셀 작성: 헤더가 2행에 있어 헤더 탐색 유연성도 확인
    wb = Workbook()
    ws = wb.active
    ws.append(["거래처 검증 시트"])  # 1행: 제목 (헤더 아님)
    ws.append(["파일폴더", "파일번호", "사업자번호", "상호명", "계약일자"])  # 2행: 헤더
    rows = [
        ["이마트24_2", "156", "123-45-67890", "주식회사 행복마트", "2021.10.08"],  # 전부 일치
        ["이마트24_2", "157", "999-99-99999", "(주)다른상호", "2020-01-01"],  # 불일치 발생
        ["이마트24_2", "158", "111-11-11111", "없는폴더상점", "2022-05-05"],  # 폴더 없음 -> 확인필요
        ["", "", "222-22-22222", "제외상점", "2023-03-03"],  # 폴더/번호 없음 -> 제외
    ]
    for r in rows:
        ws.append(r)
    xlsx_path = root / "검증대상.xlsx"
    wb.save(xlsx_path)

    # 서류 폴더 + _mock.json (더미 추출 결과)
    mocks = {
        "156": {"사업자번호": "1234567890", "상호명": "행복마트", "계약일자": "20211008"},  # 일치
        "157": {"사업자번호": "888-88-88888", "상호명": "원래상호", "계약일자": "2099.12.31"},  # 불일치
        # 158 폴더는 일부러 만들지 않음 -> 확인필요
    }
    for file_no, data in mocks.items():
        folder = work / file_no
        folder.mkdir(parents=True, exist_ok=True)
        # 더미 이미지 파일 (has_documents 통과용)
        (folder / "사업자등록증.jpg").write_bytes(b"\xff\xd8\xff\xe0dummy")
        (folder / "_mock.json").write_text(
            json.dumps(data, ensure_ascii=False), encoding="utf-8"
        )

    # ZIP 패키징
    zip_path = root / "이마트24_2.zip"
    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.write(xlsx_path, "검증대상.xlsx")
        for p in work.rglob("*"):
            zf.write(p, p.relative_to(root))
    return zip_path


def main() -> None:
    with tempfile.TemporaryDirectory() as tmp:
        root = Path(tmp)
        zip_path = build_fixture(root)
        result = verify_zip(str(zip_path), ["사업자번호", "상호명", "계약일자"])

        print("=== 검증 요약 ===")
        print(result.summary_text())
        print("결과 파일:", result.output_path)

        wb = load_workbook(result.output_path)
        ws = wb.active
        # 검증결과 열 출력
        print("\n=== 행별 결과 ===")
        for row in ws.iter_rows(min_row=2, values_only=True):
            if any(v is not None for v in row):
                print(row)

    expected_counts = {"일치": 3, "불일치": 3, "확인필요": 3, "제외": 1}
    assert result.counts == expected_counts, (
        f"기대 {expected_counts}, 실제 {result.counts}"
    )
    print("\nSMOKE TEST PASSED")


if __name__ == "__main__":
    main()
